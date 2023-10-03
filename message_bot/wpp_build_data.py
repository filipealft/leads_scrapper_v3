import database
import openpyxl
import urllib.parse
from tqdm import tqdm

def generate_message_link(phone, message):
    encoded_message = urllib.parse.quote_plus(message)
    return f'https://web.whatsapp.com/send?phone={phone}&text={encoded_message}'

def create_and_update_excel(phone, link):
    try:
        workbook = openpyxl.load_workbook("messages_sent.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Phone Number"
        sheet["B1"] = "Generated Link"
    sheet = workbook.active
    row = sheet.max_row + 1
    sheet[f"A{row}"] = phone
    sheet[f"B{row}"] = link
    workbook.save("messages_sent.xlsx")

phone_numbers = database.get_phone_leads_ready()

message_template = '''oii tudo bem? 

lucas aqui, da i9 nichos, uma empresa especializada em nichos e móveis sob medida. estava vendo o trabalho de vocês no segmento de [Segmento do Estabelecimento] e pensei que talvez gostariam de conhecer nossos produtos. se curtir a ideia, temos um desconto especial de 10% na primeira compra.

todos os nossos produtos possuem medidas customizáveis, então com certeza teremos algo que te agrade!

se quiser dar uma olhada nas nossas criações, temos um catálogo digital. basta clicar no meu número de telefone aqui no whatsapp apertar no botão "catálogo" para conferir.

dá uma espiada no nosso instagram: https://www.instagram.com/i9nichos/

estamos em florianópolis. se quiser negociar, é só chamar.
'''

# Adicionando a barra de progresso com tqdm
for phone in tqdm(phone_numbers, desc="Processing", ncols=100):
    segment = database.get_phone_and_segment_for_phone(phone)
    message = message_template.replace("[Segmento do Estabelecimento]", segment)
    generated_link = generate_message_link(phone, message)
    create_and_update_excel(phone, generated_link)
    database.update_lead_status_to_captured(phone)
